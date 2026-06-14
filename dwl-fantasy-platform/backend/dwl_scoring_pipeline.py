#!/usr/bin/env python3
"""
WWC Fantasy League - Full Pipeline
Fetches a match scorecard from ESPNcricinfo, maps players to DWL teams,
calculates fantasy points (with captain/VC multipliers), and generates
a complete Excel with per-team tabs + overall standings.

Usage:
    python dwl_scoring_pipeline.py -m 1490677 -n 1
    python dwl_scoring_pipeline.py -m 1490678 -n 2

Required file:
    WWC_Config.xlsx   -- Single Excel file with all configuration:
                        - Sheet "DWLTeams": DWL team abbreviations and names
                        - Sheet "Players": Player names, Roles, DWL teams, Country
                        - Sheet "GD", "JR", etc.: Team-specific captain/VC info
First-time setup:
    pip install playwright pandas openpyxl
    playwright install webkit
"""

import argparse
import asyncio
import io
import json
import os
import re
import sys
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from supabase import create_client, Client

# ── Import the existing fetcher ───────────────────────────────────────────────
try:
    from wwc_fantasy import WWCScorecard
except ImportError:
    print("❌ Could not import wwc_fantasy.py — make sure it's in the same directory.")
    sys.exit(1)

# Storage configuration
STORAGE_MODE = os.getenv("STORAGE_MODE", "local").lower()
BUCKET_NAME = "excel-files"
EXCEL_FILE_PATH = "DWL_Scores.xlsx"  # Path within the bucket

def is_production_mode() -> bool:
    """Check if running in production mode with Supabase storage."""
    return STORAGE_MODE == "production"


# Configure Supabase client
supabase_client: Client = None
def get_supabase_client() -> Client:
    """Initialize and return the Supabase client.  (only in production mode)."""
    global supabase_client
    if not is_production_mode():
        return None
    
    if supabase_client is None:
        url = os.getenv("SUPABASE_URL")
        key = os.getenv("SUPABASE_KEY")
        if not url or not key:
            print("⚠️ SUPABASE_URL or SUPABASE_KEY not set. Supabase features will be unavailable.")
            return None
        supabase_client = create_client(url, key)
        print("ℹ️ PROD: SUPABASE client connected. (pipeline)")
    return supabase_client
# ─────────────────────────────────────────────────────────────────────────────
# DATA STORAGE - LOCAL OR SUPABASE UPLOAD/DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────

def save_match_data(match_history: dict, data_file: str = None) -> None:
    """Save match data to Supabase (production) or local file (local) using upsert."""

    if is_production_mode():
        print("Saving match data to Supabase...")
        sb_client = get_supabase_client()
        if not sb_client:
            return
    
        try:
            # Convert the dictionary of matches into a list of dicts for Supabase
            matches_to_upsert = []
            for match_num, match_data in match_history.items():
                # We need to include the match_num as a column for the upsert to work
                match_record = {**match_data, "match_num": match_num}
                matches_to_upsert.append(match_record)
            
            # Perform upsert (update if exists, insert if new)
            # The `on_conflict="match_num"` tells PostgreSQL to update the existing row
            # if a match with the same `match_num` already exists.
            response = sb_client.table("matches").upsert(
                matches_to_upsert, 
                on_conflict="match_num"
            ).execute()
            
            if hasattr(response, 'error') and response.error:
                print(f"Supabase upsert error: {response.error}")
            else:
                print(f"✅ Saved {len(matches_to_upsert)} match entries to Supabase")
        except Exception as e:
            print(f"Error saving match data to Supabase: {e}")
    else:
        # Local mode - save to JSON file
        print("Saving match data to Local file:...")
        print(f"   Path: {os.path.abspath(data_file)}")
        if data_file:
            serializable = {}
            for mn, match_data in match_history.items():
                serializable[str(mn)] = {
                    "match_entry": match_data.get("match_entry", {}),
                    "team1_country": match_data.get("team1_country", ""),
                    "team2_country": match_data.get("team2_country", ""),
                    "match_winner": match_data.get("match_winner", ""),
                    "match_title": match_data.get("match_title", "")
                }
            with open(data_file, 'w') as f:
                json.dump(serializable, f, indent=2)


def load_match_data(data_file: str = "wwc_match_data.json") -> dict:
    """Load match data from Supabase (production) or local JSON file (local)."""
    if is_production_mode():
        # Production mode - read from Supabase table

        print("Loading match data from Supabase...")
        sb_client = get_supabase_client()
        if not sb_client:
            return {}
    
        try:
            response = sb_client.table("matches").select("*").execute()
            if response.data:
                # Convert the list of matches from Supabase back to the dictionary format
                # your app expects: { match_num: { ...match_data } }
                match_dict = {}
                for match in response.data:
                    match_num = match.pop("match_num")
                    match_dict[match_num] = match
                return match_dict
            return {}
        except Exception as e:
            print(f"Error loading match data from Supabase: {e}")
            return {}
    else:
        # Local mode - read from JSON file

        print("Loading match data from Local file:...")
        print(f"   Path: {os.path.abspath(data_file)}")
        if Path(data_file).exists():
            with open(data_file, 'r') as f:
                data = json.load(f)
            match_history = {}
            for mn_str, match_data in data.items():
                mn = int(mn_str)
                match_history[mn] = {
                    "match_entry": match_data.get("match_entry", {}),
                    "team1_country": match_data.get("team1_country", ""),
                    "team2_country": match_data.get("team2_country", ""),
                    "match_winner": match_data.get("match_winner", ""),
                    "match_title": match_data.get("match_title", "")
                }
            return match_history
        return {}


# ─────────────────────────────────────────────────────────────────────────────
# NAME MATCHING - SIMPLIFIED USING ESPN FIELDS
# ─────────────────────────────────────────────────────────────────────────────

def _normalise(name: str) -> str:
    """Lowercase, strip accents-ish, collapse whitespace."""
    if not name:
        return ""
    return re.sub(r"\s+", " ", name.strip().lower())


def build_dwl_lookup(roster_names: list[str]) -> dict[str, str]:
    """
    Build lookup dict mapping normalized DWL player names to canonical names.
    Also builds variations for matching.
    """
    lookup = {}
    
    for canonical in roster_names:
        normalized = _normalise(canonical)
        if normalized:
            lookup[normalized] = canonical
            
            # Add first name + last name variation
            parts = canonical.split()
            if len(parts) >= 2:
                first_last = f"{parts[0]} {parts[-1]}"
                lookup[_normalise(first_last)] = canonical
    
    return lookup


def match_player_to_roster(
    espn_player: dict,
    dwl_lookup: dict[str, str],
    player_team: dict[str, str],
    player_country_mapping: dict[str, str] = None,
) -> str | None:
    """
    Match ESPN player to DWL roster using ESPN's longName and fieldingName.
    
    ESPN player dict contains:
    - longName: Full name (e.g., "Beth Mooney")
    - fieldingName: Last name for fielding (e.g., "Mooney")
    - short_name: ESPN short name (e.g., "BL Mooney")
    - team: Country abbreviation (e.g., "AUS-W")
    """
    if not espn_player:
        return None
    
    long_name = espn_player.get("full_name", "")  # This maps to longName from ESPN
    short_name = espn_player.get("short_name", "")
    fielding_name = espn_player.get("fielding_name", "")  # We'll add this to the fetcher
    country = espn_player.get("team", "")
    
    if not long_name:
        return None
    
    # Normalize the long name for lookup
    normalized_long = _normalise(long_name)
    
    # Strategy 1: Direct match on full name (highest confidence)
    if normalized_long in dwl_lookup:
        candidate = dwl_lookup[normalized_long]
        # Verify country matches if we have mapping
        if player_country_mapping:
            expected_country = player_country_mapping.get(candidate, "")
            if expected_country and expected_country != country:
                # Country mismatch - don't use this match
                pass
            else:
                return candidate
    
    # Strategy 2: Match on short name (e.g., "BL Mooney")
    if short_name:
        normalized_short = _normalise(short_name)
        if normalized_short in dwl_lookup:
            candidate = dwl_lookup[normalized_short]
            if player_country_mapping:
                expected_country = player_country_mapping.get(candidate, "")
                if expected_country and expected_country != country:
                    pass
                else:
                    return candidate
    
    # Strategy 3: Match on fielding name (last name) - but only if no duplicates in team
    if fielding_name and country:
        normalized_fielding = _normalise(fielding_name)
        
        # Find all DWL players from this country with matching last name
        matching_players = []
        for dwl_player, dwl_country in player_country_mapping.items():
            if dwl_country == country:
                dwl_last = _normalise(dwl_player.split()[-1] if dwl_player.split() else "")
                if dwl_last == normalized_fielding:
                    matching_players.append(dwl_player)
        
        # If exactly one match, use it
        if len(matching_players) == 1:
            return matching_players[0]
    
    # Strategy 4: Try to match by checking if long name contains DWL name or vice versa
    for dwl_name in dwl_lookup.values():
        normalized_dwl = _normalise(dwl_name)
        
        # Check if long name contains DWL name or DWL name contains long name
        if normalized_long in normalized_dwl or normalized_dwl in normalized_long:
            if player_country_mapping:
                expected_country = player_country_mapping.get(dwl_name, "")
                if expected_country and expected_country != country:
                    continue
            return dwl_name
    
    return None


# ─────────────────────────────────────────────────────────────────────────────
# READ CONFIG FILES
# ─────────────────────────────────────────────────────────────────────────────

def read_dwl_config(config_path: str, debug: bool = False):
    """
    Read all DWL configuration from a single Excel file.
    
    Expected sheets:
    - "DWLTeams": Columns: Abbrv, Team Name
    - "Players": Columns: Player Name, Role, DWL Team, Country
    - Team sheets (e.g., "GD", "JR", etc.): Captain/VC info with specific format
    """
    warnings.simplefilter(action='ignore', category=UserWarning)
    
    CAPTAIN_MULT = 2.0
    VC_MULT = 1.5

    # Teams
    teams_df = pd.read_excel(config_path, sheet_name="DWLTeams")
    teams_abbr = dict(zip(teams_df["Abbrv"].str.strip(), teams_df["Team Name"].str.strip()))
    abbr_to_name = dict(teams_abbr)
    name_to_abbr = {v: k for k, v in abbr_to_name.items()}

    # Map full country names to ESPN abbreviations
    country_name_to_abbr = {
        "Australia": "AUS-W",
        "Bangladesh": "BAN-W",
        "India": "IND-W",
        "Netherlands": "NL-W",
        "Pakistan": "PAK-W",
        "South Africa": "SA-W",
        "England": "ENG-W",
        "Ireland": "IRE-W",
        "New Zealand": "NZ-W",
        "Scotland": "SCO-W",
        "Sri Lanka": "SL-W",
        "West Indies": "WI-W",
    }

    # Rosters
    players_df = pd.read_excel(config_path, sheet_name="Players")
    players_df.columns = [c.strip() for c in players_df.columns]
    rosters: dict[str, list[str]] = {}
    player_team: dict[str, str] = {}
    player_country: dict[str, str] = {}
    player_role: dict[str, str] = {}
    player_price: dict[str, float] = {}
    
    # Track injured/replaced players
    player_status: dict[str, str] = {}  # "Injured" or "Replacement"
    player_replacement_map: dict[str, str] = {}  # injured_player -> replacement_player
    
    # Extract wicketkeepers from Players sheet based on Role column
    country_wicketkeepers: dict[str, list[str]] = {}
    
    print(f"\n📋 Loading players from Excel...")
    
    for _, row in players_df.iterrows():
        pname = str(row["Player Name"]).strip()
        team = str(row["DWL Team"]).strip()
        country_full = str(row.get("Country", "")).strip()
        role = str(row.get("Role", "")).strip().lower()
        
        # Check for Replaced column (could be "Replaced" or "Status" or similar)
        replaced_value = str(row.get("Replaced", "")).strip() if "Replaced" in row else ""
        if not replaced_value and "Status" in row:
            replaced_value = str(row.get("Status", "")).strip()
        
        if pname and pname != "nan" and team and team != "nan":
            rosters.setdefault(team, []).append(pname)
            player_team[pname] = team
            player_role[pname] = role
            
            # Track player status
            if replaced_value and replaced_value != "nan":
                player_status[pname] = replaced_value
                if debug:
                    print(f"   {pname} status: {replaced_value}")
            
            if country_full and country_full != "nan":
                country_abbr = country_name_to_abbr.get(country_full)
                if not country_abbr:
                    for full_name, abbr in country_name_to_abbr.items():
                        if full_name.lower() in country_full.lower():
                            country_abbr = abbr
                            break
                if country_abbr:
                    player_country[pname] = country_abbr
                    if debug:
                        print(f"   Mapped {pname} -> DWL: {team}, Country: {country_abbr} (from '{country_full}')")
            
            # Build wicketkeeper list by country
            if role == "wicketkeeper" and country_full and country_full != "nan":
                country_abbr = country_name_to_abbr.get(country_full)
                if country_abbr:
                    country_wicketkeepers.setdefault(country_abbr, []).append(pname)

            price = row.get("Sold Price", 0) or row.get("Price", 0)
            if price and price != "nan":
                player_price[pname] = float(price) if isinstance(price, (int, float)) else 0

    if debug:
        print(f"\n📋 All players loaded with countries:")
        for pname in sorted(player_team.keys()):
            country = player_country.get(pname, "NOT SET")
            dwl = player_team.get(pname, "?")
            role = player_role.get(pname, "?")
            status = player_status.get(pname, "ACTIVE")
            print(f"   {pname:<30} -> DWL: {dwl:<20} Country: {country} Role: {role} Status: {status}")
        
        print(f"\n📋 Wicketkeepers by country:")
        for country, wks in country_wicketkeepers.items():
            print(f"   {country}: {', '.join(wks)}")

    all_roster_names = list(player_team.keys())
    dwl_lookup = build_dwl_lookup(all_roster_names)

    # Captain / VC from team-specific sheets
    capt_vc: dict[str, dict[str, float]] = {}
    
    # Get all sheet names in the Excel file
    xl = pd.ExcelFile(config_path)
    sheet_names = xl.sheet_names
    
    # Team sheets are those that match DWL team abbreviations
    dwl_abbrs = set(teams_abbr.keys())
    
    for abbr in dwl_abbrs:
        if abbr in sheet_names:
            try:
                raw = pd.read_excel(config_path, sheet_name=abbr, header=None)
                caps = {}
                for _, row in raw.iloc[5:].iterrows():
                    pname = str(row.iloc[1]).strip()
                    role = str(row.iloc[3]).strip().lower()
                    if pname and pname != "nan":
                        role_lower = role.lower()  # Convert to lowercase for comparison
                        if "captain" in role_lower and "vice" not in role_lower:
                            caps[pname] = CAPTAIN_MULT
                        elif "vice" in role_lower:
                            caps[pname] = VC_MULT
                capt_vc[abbr] = caps
                if debug:
                    print(f"   Loaded captain/VC for {abbr}: {len(caps)} players")
            except Exception as e:
                if debug:
                    print(f"   Warning: Could not load sheet '{abbr}': {e}")
        else:
            if debug:
                print(f"   No sheet found for team {abbr}")

    return (
        teams_abbr, rosters, player_team, capt_vc,
        dwl_lookup, abbr_to_name, name_to_abbr, all_roster_names,
        player_country, country_wicketkeepers, player_status,
    )


# ─────────────────────────────────────────────────────────────────────────────
# PROCESS FETCHED SCORECARD
# ─────────────────────────────────────────────────────────────────────────────

def process_scorecard(
    scorecard_data: dict,
    match_num: int,
    dwl_lookup: dict,
    player_team: dict,
    capt_vc: dict,
    name_to_abbr: dict,
    abbr_to_name: dict,
    player_country: dict = None,
    country_wicketkeepers: dict = None,
    debug: bool = False,
):
    """Convert fetched scorecard into match history entry."""
    info = scorecard_data["match_info"]
    players = scorecard_data["players"]
    match_winner = scorecard_data["match_info"]["winner"]
    
    # Extract actual country names from the match
    match_title = info.get('title', '')
    home_team = info.get('home_team', '')
    away_team = info.get('away_team', '')
    
    # Get country names from the scorecard players
    countries_in_match = set()
    for p in players:
        country = p.get("team", "")
        if country:
            countries_in_match.add(country)
    
    countries_list = sorted(list(countries_in_match))
    team1_country = countries_list[0] if len(countries_list) > 0 else ""
    team2_country = countries_list[1] if len(countries_list) > 1 else ""
        
    match_entry = {}
    unmatched = []

    for p in players:
        full_name = p.get("full_name", "")
        short_name = p.get("short_name", "")
        fielding_name = p.get("fielding_name", "")
        country = p.get("team", "")
        
        # Create ESPN player dict for matching
        espn_player = {
            "full_name": full_name,
            "short_name": short_name,
            "fielding_name": fielding_name,
            "team": country,
        }
        
        # Regular match stats
        bat_list = [b for b in p.get("batting", []) if b.get("runs") is not None]
        runs = sum(b["runs"] for b in bat_list)
        balls = sum(b.get("balls", 0) or 0 for b in bat_list)
        fours = sum(b.get("fours", 0) or 0 for b in bat_list)
        sixes = sum(b.get("sixes", 0) or 0 for b in bat_list)
        is_out = any(b.get("is_out", False) for b in bat_list)
        dismissal = bat_list[-1].get("dismissal", "") if bat_list else ""

        bowl_list = [b for b in p.get("bowling", []) if b.get("overs") is not None]
        overs = sum(b.get("overs", 0) or 0 for b in bowl_list)
        maidens = sum(b.get("maidens", 0) or 0 for b in bowl_list)
        runs_c = sum(b.get("runs", 0) or 0 for b in bowl_list)
        wickets = sum(b.get("wickets", 0) or 0 for b in bowl_list)
        has_hattrick = any(b.get("has_hattrick", False) for b in bowl_list)

        fielding = p.get("fielding", {})
        catches = fielding.get("catches", 0)
        run_outs = fielding.get("run_outs", 0)
        stumpings = fielding.get("stumpings", 0)

        # Super Over stats (if present)
        so_bat_list = [b for b in p.get("super_over_batting", []) if b.get("runs") is not None]
        so_runs = sum(b["runs"] for b in so_bat_list)
        so_balls = sum(b.get("balls", 0) or 0 for b in so_bat_list)
        so_fours = sum(b.get("fours", 0) or 0 for b in so_bat_list)
        so_sixes = sum(b.get("sixes", 0) or 0 for b in so_bat_list)
        so_is_out = any(b.get("is_out", False) for b in so_bat_list)

        so_bowl_list = [b for b in p.get("super_over_bowling", []) if b.get("overs") is not None]
        so_overs = sum(b.get("overs", 0) or 0 for b in so_bowl_list)
        so_maidens = sum(b.get("maidens", 0) or 0 for b in so_bowl_list)
        so_runs_c = sum(b.get("runs", 0) or 0 for b in so_bowl_list)
        so_wickets = sum(b.get("wickets", 0) or 0 for b in so_bowl_list)

        so_fielding = p.get("super_over_fielding", {})
        so_catches = so_fielding.get("catches", 0)
        so_run_outs = so_fielding.get("run_outs", 0)
        so_stumpings = so_fielding.get("stumpings", 0)

        is_motm = p.get("is_motm", False)

        # Check if player actually played
        has_batting_contribution = runs > 0 or balls > 0 or fours > 0 or sixes > 0 or is_out
        has_bowling_contribution = overs > 0 or maidens > 0 or runs_c > 0 or wickets > 0
        has_fielding_contribution = catches > 0 or run_outs > 0 or stumpings > 0
        has_so_contribution = (so_runs > 0 or so_balls > 0 or so_fours > 0 or so_sixes > 0 or
                               so_is_out or so_overs > 0 or so_maidens > 0 or so_runs_c > 0 or
                               so_wickets > 0 or so_catches > 0 or so_run_outs > 0 or so_stumpings > 0)
        is_motm_bonus = is_motm
        
        # Skip players with absolutely no contribution
        if not (has_batting_contribution or has_bowling_contribution or
                has_fielding_contribution or has_so_contribution or is_motm_bonus):
            if debug:
                print(f"DEBUG: Skipping {full_name} ({short_name}) - no contribution")
            continue

        # Use category points from the fetcher if available
        reg_batting_pts = p.get("reg_batting", 0)
        reg_bowling_pts = p.get("reg_bowling", 0)
        reg_fielding_pts = p.get("reg_fielding", 0)

        so_batting_pts = p.get("so_batting", 0)
        so_bowling_pts = p.get("so_bowling", 0)
        so_fielding_pts = p.get("so_fielding", 0)

        # Use pre-calculated points from the fetcher
        raw_pts = p.get("fantasy_points", 0)
        regular_pts = p.get("regular_points", 0)
        so_pts = p.get("super_over_points", 0)

        # Map to DWL roster using simplified matching
        canon = match_player_to_roster(
            espn_player, dwl_lookup, player_team, player_country
        )
        
        if debug:
            if canon:
                print(f"   ✅ MATCH: '{short_name}' ({full_name}) → '{canon}' (Country: {country})")
            else:
                print(f"   ❌ NO MATCH: '{short_name}' ({full_name}) (Country: {country})")
        
        if canon is None:
            unmatched.append(f"{full_name} ({short_name}) - {country}")
            continue

        team_name = player_team.get(canon, "")
        if not team_name:
            unmatched.append(f"{canon} [no DWL team]")
            continue

        abbr = name_to_abbr.get(team_name, "")
        mult = capt_vc.get(abbr, {}).get(canon, 1.0)

        match_entry[canon] = {
            "pts": raw_pts,
            "regular_pts": regular_pts,
            "regular_batting_pts": reg_batting_pts,
            "regular_bowling_pts": reg_bowling_pts,
            "regular_fielding_pts": reg_fielding_pts,
            "super_over_pts": so_pts,
            "so_batting_pts": so_batting_pts,
            "so_bowling_pts": so_bowling_pts,
            "so_fielding_pts": so_fielding_pts,
            "final_pts": raw_pts * mult,
            "is_motm": is_motm,
            "dwl_team": team_name,
            "country": country,
            "_stats": dict(
                runs=runs, balls=balls, fours=fours, sixes=sixes,
                overs=overs, maidens=maidens, runs_c=runs_c, wickets=wickets, has_hattrick=has_hattrick,
                catches=catches, run_outs=run_outs, stumpings=stumpings,
                so_runs=so_runs, so_balls=so_balls, so_fours=so_fours, so_sixes=so_sixes,
                so_overs=so_overs, so_maidens=so_maidens, so_runs_c=so_runs_c, so_wickets=so_wickets,
                so_catches=so_catches, so_run_outs=so_run_outs, so_stumpings=so_stumpings,
                mult=mult,
            ),
        }

    match_info_str = (
        f"M{match_num}: {info.get('title','')}"
        f"  |  {info.get('date','')}"
        f"  |  {info.get('result','')}"
        + (f"  |  PoM: {info.get('player_of_match','')}" if info.get("player_of_match") else "")
        + (f"  |  ⚡ SUPER OVER" if info.get('has_super_over', False) else "")
    )
    
    # Return match entry with country metadata
    return {
        "match_entry": match_entry,
        "team1_country": team1_country,
        "team2_country": team2_country,
        "match_winner": match_winner,
        "match_title": match_title
    }, unmatched, match_info_str


# ─────────────────────────────────────────────────────────────────────────────
# PRINT MATCH SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def print_match_summary(match_entry, teams_abbr, name_to_abbr, player_team):
    """Print per-DWL-team summary for the match."""
    by_team: dict[str, list] = {}
    for pname, pdata in match_entry.items():
        team = pdata["dwl_team"]
        by_team.setdefault(team, []).append((pname, pdata))

    print()
    for team_name in sorted(by_team.keys()):
        abbr = name_to_abbr.get(team_name, "??")
        players = sorted(by_team[team_name], key=lambda x: -x[1]["final_pts"])
        team_total = sum(p[1]["final_pts"] for p in players)
        print(f"  [{abbr}] {team_name}  —  {team_total:+.1f} pts")
        for pname, pd in players:
            s = pd["_stats"]
            mult = s["mult"]
            raw = pd["pts"]
            final = pd["final_pts"]
            motm = " ⭐PoM" if pd["is_motm"] else ""
            mult_str = f" (×{mult})" if mult > 1 else ""
            print(f"        {pname:<28} {final:>6.0f} pts  [raw {raw}{mult_str}]{motm}")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATION - STYLES
# ─────────────────────────────────────────────────────────────────────────────

HEADER_BG = "1F3864"
SUBHDR_BG = "2E75B6"
CAP_BG    = "FFD700"
VC_BG     = "C0C0C0"
ALT_BG    = "DEEAF1"
TOTAL_BG  = "E2EFDA"
MOTM_FG   = "FF6600"

_thin = Side(style="thin", color="CCCCCC")
TB = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _fill(h): return PatternFill("solid", fgColor=h)
def _center(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def _left(): return Alignment(horizontal="left", vertical="center")
def _font(size=10, bold=False, color="000000", italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def _shdr(cell, text, bg=SUBHDR_BG):
    cell.value = text
    cell.font = _font(10, True, "FFFFFF")
    cell.fill = _fill(bg)
    cell.alignment = _center(wrap=True)
    cell.border = TB

def _hdr(cell, text, bg=HEADER_BG):
    cell.value = text
    cell.font = _font(11, True, "FFFFFF")
    cell.fill = _fill(bg)
    cell.alignment = _left()
    cell.border = TB


def _init_team_sheet(ws, team_name, abbr, roster, capt_vc_team, player_country, player_status=None):
    """Initialize a new team sheet with headers."""
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    c = ws.cell(row=1, column=1, value=f"{team_name.upper()}  ({abbr})")
    c.font = _font(14, True, "FFFFFF")
    c.fill = _fill(HEADER_BG)
    c.alignment = _center()
    
    # Legend
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    legend_text = "🏆 Captain (2×)   |   🥈 Vice-Captain (1.5×)   |   ⭐ Player of the Match"
    if player_status:
        legend_text += "   |   🏥 Injured   |   🔄 Replacement"
    c = ws.cell(row=2, column=1, value=legend_text)
    c.font = _font(9, False, "555555", italic=True)
    c.alignment = _center()
    
    # Headers (#, Player, Country, ×, Total)
    _shdr(ws.cell(row=3, column=1), "#")
    _shdr(ws.cell(row=3, column=2), "Player")
    _shdr(ws.cell(row=3, column=3), "Country")
    _shdr(ws.cell(row=3, column=4), "×")
    _shdr(ws.cell(row=3, column=5), "Total")
    
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28  # Increased to accommodate status indicators
    ws.column_dimensions["C"].width = 15  # Increased width for full country names
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 9
    
    # Country name mapping
    country_full_names = {
        'AUS-W': 'Australia',
        'BAN-W': 'Bangladesh',
        'ENG-W': 'England',
        'IND-W': 'India',
        'IRE-W': 'Ireland',
        'NL-W': 'Netherlands',
        'NZ-W': 'New Zealand',
        'PAK-W': 'Pakistan',
        'SCO-W': 'Scotland',
        'SA-W': 'South Africa',
        'SL-W': 'Sri Lanka',
        'WI-W': 'West Indies',
    }
    
    # Sort players alphabetically by name
    sorted_roster = sorted(roster)
    
    # Player rows
    for ri, pname in enumerate(sorted_roster):
        row = 4 + ri
        mult = capt_vc_team.get(pname, 1.0)
        country = player_country.get(pname, "")
        
        # Get player status if available
        status = player_status.get(pname, "") if player_status else ""
        
        # Convert country abbreviation to full name
        country_display = country_full_names.get(country, country)
        
        if mult == 2.0:
            bg = CAP_BG
        elif mult == 1.5:
            bg = VC_BG
        elif ri % 2 == 1:
            bg = ALT_BG
        else:
            bg = "FFFFFF"
        
        c = ws.cell(row=row, column=1, value=ri + 1)
        c.font = _font(10); c.fill = _fill(bg); c.alignment = _center(); c.border = TB
        
        # Add status indicator to player name
        prefix = ""
        if mult == 2.0:
            prefix = "🏆 "
        elif mult == 1.5:
            prefix = "🥈 "
        
        # Add injury/replacement indicator
        status_suffix = ""
        if status == "Injured":
            status_suffix = " 🏥"
        elif status == "Replacement":
            status_suffix = " 🔄"
        
        c = ws.cell(row=row, column=2, value=f"{prefix}{pname}{status_suffix}")
        c.font = _font(10, bold=(mult > 1 or status))
        if status == "Injured":
            c.font = _font(10, bold=True, italic=True, color="FF0000")
        elif status == "Replacement":
            c.font = _font(10, bold=True, color="008000")
        c.fill = _fill(bg); c.alignment = _left(); c.border = TB
        
        c = ws.cell(row=row, column=3, value=country_display)
        c.font = _font(10); c.fill = _fill(bg); c.alignment = _left(); c.border = TB
        
        c = ws.cell(row=row, column=4, value=f"{mult:.1f}×")
        c.font = _font(10, bold=(mult > 1)); c.fill = _fill(bg); c.alignment = _center(); c.border = TB
        
        # Total starts at 0
        c = ws.cell(row=row, column=5, value=0)
        c.font = _font(10, bold=True); c.fill = _fill(TOTAL_BG if mult == 1.0 else bg)
        c.alignment = _center(); c.border = TB
    
    # Team total row
    trow = 4 + len(sorted_roster)
    ws.cell(row=trow, column=1).value = ""
    _hdr(ws.cell(row=trow, column=2), "TEAM TOTAL")
    ws.cell(row=trow, column=3).value = ""
    ws.cell(row=trow, column=4).value = ""
    ws.cell(row=trow, column=4).alignment = _center()
    c = ws.cell(row=trow, column=5, value=0)
    c.font = _font(11, True, "FFFFFF"); c.fill = _fill(HEADER_BG)
    c.alignment = _center(); c.border = TB
    
    ws.freeze_panes = "E4"


def write_team_sheet(ws, team_name, abbr, roster, capt_vc_team, match_history, all_match_nums, player_country):
    """Write complete team sheet from match_history."""
    # Clear existing columns beyond column 5
    if ws.max_column > 5:
        for col_idx in range(ws.max_column, 5, -1):
            ws.delete_cols(col_idx)
    
    # Add match columns starting at column 6
    for idx, mn in enumerate(all_match_nums):
        col_idx = 6 + idx
        _shdr(ws.cell(row=3, column=col_idx), f"M{mn}")
        ws.column_dimensions[get_column_letter(col_idx)].width = 7
    
    # Track per-match totals
    per_match_totals = {mn: 0.0 for mn in all_match_nums}
    
    # Sort players alphabetically
    sorted_roster = sorted(roster)
    
    # Update each player row
    for ri, pname in enumerate(sorted_roster):
        row = 4 + ri
        mult = capt_vc_team.get(pname, 1.0)
        
        if mult == 2.0:
            bg = CAP_BG
        elif mult == 1.5:
            bg = VC_BG
        elif ri % 2 == 1:
            bg = ALT_BG
        else:
            bg = "FFFFFF"
        
        player_total = 0.0
        
        # Fill match columns
        for idx, mn in enumerate(all_match_nums):
            col_idx = 6 + idx
            cell = ws.cell(row=row, column=col_idx)
            match_data = match_history.get(mn, {})
            match_entry = match_data.get("match_entry", {}) if isinstance(match_data, dict) else {}
            mdata = match_entry.get(pname, {}) if isinstance(match_entry, dict) else {}
            
            if mdata and mdata.get("dwl_team") == team_name:
                val = mdata.get("final_pts", 0)
                val_disp = round(val, 1) if val != int(val) else int(val)
                cell.value = val_disp
                is_motm = mdata.get("is_motm", False)
                cell.font = _font(10, bold=is_motm, color=(MOTM_FG if is_motm else "000000"))
                player_total += val
                per_match_totals[mn] += val
            else:
                cell.value = "-"
                cell.font = _font(10, color="AAAAAA")
            
            cell.fill = _fill(bg)
            cell.alignment = _center()
            cell.border = TB
        
        # Update total column
        total_cell = ws.cell(row=row, column=5)
        pt_disp = round(player_total, 1) if player_total != int(player_total) else int(player_total)
        total_cell.value = pt_disp
        total_cell.font = _font(10, bold=True)
        total_cell.fill = _fill(TOTAL_BG if mult == 1.0 else bg)
        total_cell.alignment = _center()
        total_cell.border = TB
    
    # Team total row
    trow = 4 + len(sorted_roster)
    
    # Calculate grand total
    grand_total = 0.0
    for ri in range(len(sorted_roster)):
        row = 4 + ri
        total_cell = ws.cell(row=row, column=5)
        if isinstance(total_cell.value, (int, float)):
            grand_total += total_cell.value
    
    team_total_cell = ws.cell(row=trow, column=5)
    team_total_cell.value = round(grand_total, 1) if grand_total != int(grand_total) else int(grand_total)
    
    # Update per-match totals in team total row
    for idx, mn in enumerate(all_match_nums):
        col_idx = 6 + idx
        cell = ws.cell(row=trow, column=col_idx)
        val = per_match_totals.get(mn, 0)
        val_disp = round(val, 1) if val != int(val) else int(val)
        cell.value = val_disp if val_disp else 0
        cell.font = _font(10, True, "FFFFFF")
        cell.fill = _fill(HEADER_BG)
        cell.alignment = _center()
        cell.border = TB
    
    return grand_total, per_match_totals


def write_standings_sheet(ws, standings, all_match_nums):
    """Write the Standings sheet."""
    standings_sorted = sorted(standings, key=lambda x: -x[2])
    active = sorted(all_match_nums)
    
    ncols = 3 + len(active)
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="🏏  DWL FANTASY LEAGUE  —  OVERALL STANDINGS")
    c.font = _font(14, True, "FFFFFF"); c.fill = _fill(HEADER_BG); c.alignment = _center()
    
    # Headers
    hdrs = ["POS", "TEAM", "TOTAL"] + [f"M{mn}" for mn in active]
    for ci, h in enumerate(hdrs, 1):
        _shdr(ws.cell(row=2, column=ci), h)
    
    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 9
    for i in range(len(active)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 7
    
    MEDAL = {1: "🥇", 2: "🥈", 3: "🥉"}
    POS_BG = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}
    
    for ri, (team_name, abbr, total, mpts) in enumerate(standings_sorted):
        row = 3 + ri
        pos = ri + 1
        bg = POS_BG.get(pos, "FFFFFF" if ri % 2 == 0 else ALT_BG)
        
        c = ws.cell(row=row, column=1, value=MEDAL.get(pos, str(pos)))
        c.font = _font(11, True); c.fill = _fill(bg); c.alignment = _center(); c.border = TB
        
        c = ws.cell(row=row, column=2, value=f"{team_name}  [{abbr}]")
        c.font = _font(10, bold=(pos <= 3)); c.fill = _fill(bg); c.alignment = _left(); c.border = TB
        
        t_disp = round(total, 1) if total != int(total) else int(total)
        c = ws.cell(row=row, column=3, value=t_disp)
        c.font = _font(11, True); c.fill = _fill(TOTAL_BG); c.alignment = _center(); c.border = TB
        
        for ci, mn in enumerate(active):
            val = mpts.get(mn, 0)
            val_disp = round(val, 1) if val != int(val) else int(val)
            c = ws.cell(row=row, column=4 + ci)
            c.value = val_disp if val_disp else "-"
            c.font = _font(10); c.fill = _fill(bg); c.alignment = _center(); c.border = TB
    
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 26


def generate_excel(
    sb_client: Client | None,  # Supabase client for storage operations
    output_path: str,
    teams_abbr: dict,
    rosters: dict,
    capt_vc: dict,
    match_history: dict,
    all_match_nums: list,
    player_country: dict,
    player_status: dict = None,
    use_supabase: bool = False,
):
    """Generate or update Excel file - saves locally or to Supabase based on mode."""
    wb = Workbook()
    wb.remove(wb.active)
    
    standings = []
    
    for abbr in sorted(teams_abbr.keys()):
        team_name = teams_abbr[abbr]
        roster = rosters.get(team_name, [])
        if not roster:
            continue
        
        ws = wb.create_sheet(title=abbr)
        _init_team_sheet(ws, team_name, abbr, roster, capt_vc.get(abbr, {}), player_country, player_status)
        grand_total, per_match = write_team_sheet(
            ws, team_name, abbr, roster, capt_vc.get(abbr, {}),
            match_history, all_match_nums, player_country
        )
        standings.append((team_name, abbr, grand_total, per_match))
    
    ws_s = wb.create_sheet(title="Standings", index=0)
    write_standings_sheet(ws_s, standings, all_match_nums)
    
    if use_supabase and is_production_mode():
        # Save to Supabase Storage
        try:
            # Save to bytes instead of disk
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)

            if sb_client:
                # First, try to remove existing file if it exists
                try:
                    sb_client.storage.from_(BUCKET_NAME).remove([EXCEL_FILE_PATH])
                    print(f"Removed existing file: {EXCEL_FILE_PATH}")
                except:
                    pass  # File doesn't exist, that's fine
    
                # Upload to Supabase Storage
                sb_client.storage.from_(BUCKET_NAME).upload(
                    EXCEL_FILE_PATH, 
                    excel_bytes.getvalue(),
                )
                print(f"✅ Excel file saved to Supabase Storage: {BUCKET_NAME}/{EXCEL_FILE_PATH}")
            else:
                # Fallback to local
                wb.save(output_path)
                print(f"⚠️ Supabase unavailable - saved locally as {os.path.abspath(output_path)}")
        except Exception as e:
            print(f"❌ Error uploading to Supabase: {e}")
            # Fallback: save locally for development
            local_path = "DWL_Scores.xlsx"
            wb.save(local_path)
            print(f"⚠️ Saved locally as {local_path}")
    else:
        # Save locally
        wb.save(output_path)
        abs_path = os.path.abspath(output_path)
        print(f"✅ Excel file saved locally: {abs_path}")
    
    return standings


def extract_espn_names(scorecard_data: dict, output_file: str = "espn_names.txt"):
    """Extract all player names from ESPN scorecard for analysis."""
    players = scorecard_data.get("players", [])
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("ESPN Player Names Extraction\n")
        f.write("=" * 60 + "\n\n")
        
        for p in players:
            full = p.get("full_name", "N/A")
            short = p.get("short_name", "N/A")
            fielding = p.get("fielding_name", "N/A")
            team = p.get("team", "N/A")
            
            f.write(f"Full Name: {full}\n")
            f.write(f"Short Name: {short}\n")
            f.write(f"Fielding Name: {fielding}\n")
            f.write(f"Team: {team}\n")
            f.write("-" * 40 + "\n")
    
    print(f"✅ ESPN names saved to {output_file}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

async def run(args):
    config_file = args.config
    output_file = args.output
    match_id = args.match_id
    match_num = args.match_num
    data_file = "wwc_match_data.json"

    if not Path(config_file).exists():
        print(f"❌ Config file not found: {config_file}")
        sys.exit(1)

    print(f"📖 Loading DWL config from {config_file} ...")
    (
        teams_abbr, rosters, player_team, capt_vc,
        dwl_lookup, abbr_to_name, name_to_abbr, all_roster_names,
        player_country, country_wicketkeepers, player_status,
    ) = read_dwl_config(config_file, debug=args.debug)

    # Load existing match data from Supabase
    match_history = load_match_data(data_file)
    all_match_nums = sorted(match_history.keys())
    
    print(f"\n   Teams: {', '.join(sorted(teams_abbr.keys()))}")
    print(f"   Existing matches in history: {len(match_history)}")
    if match_history:
        print(f"   Match numbers: {all_match_nums}")

    if match_id:
        if match_num in match_history and not args.force:
            print(f"\n⚠️ M{match_num} already exists. Use --force to overwrite.\n")
        else:
            print(f"\n📡 Fetching ESPNcricinfo match {match_id} → M{match_num} ...")
            sc = WWCScorecard(match_id, debug=args.debug)
            await sc.fetch()

            if not args.quiet:
                sc.print_pretty_scorecard()

            data = sc.get_full_scorecard()
            
            try:
                os.makedirs("espn_names_match", exist_ok=True)
                # Extract ESPN names for analysis
                extract_espn_names(data, f"espn_names_match/match_{match_num}.txt")
            except Exception as espn_error:
                print(f"⚠️ Warning: Could not save ESPN names match: {espn_error}")
            
            result, unmatched, info_str = process_scorecard(
                data, match_num, dwl_lookup, player_team, capt_vc,
                name_to_abbr, abbr_to_name,
                player_country=player_country,
                country_wicketkeepers=country_wicketkeepers,
                debug=args.debug,
            )
            
            match_entry = result["match_entry"]
            team1_country = result["team1_country"]
            team2_country = result["team2_country"]
            match_winner = result["match_winner"]
            match_title = result["match_title"]

            print(f"\n✅ {info_str}")
            print(f"   Matched {len(match_entry)} players to DWL rosters.")
            print(f"   Countries: {team1_country} vs {team2_country}")

            if unmatched:
                print(f"\n⚠️ {len(unmatched)} players NOT matched to any DWL roster:")
                for u in unmatched[:10]:
                    print(f"        {u}")
                if len(unmatched) > 10:
                    print(f"        ... and {len(unmatched) - 10} more")

            print("\n📊 DWL points breakdown for this match:")
            print_match_summary(match_entry, teams_abbr, name_to_abbr, player_team)

            # Update history and save to Supabase
            match_history[match_num] = {
                "match_entry": match_entry,
                "team1_country": team1_country,
                "team2_country": team2_country,
                "match_winner": match_winner,
                "match_title": match_title
            }
            all_match_nums = sorted(match_history.keys())
            save_match_data(match_history, data_file)
            
            print(f"   Updated match numbers: {all_match_nums}")
    else:
        print("\nℹ️ No --match-id given — regenerating Excel from existing history only.")
        if not match_history:
            print("   No existing matches found. Nothing to generate.")
            return

    print(f"\n📝 Writing Excel → {output_file} ...")

    # Determine storage mode
    use_supabase = is_production_mode()
    print(f"   Storage mode: {'PRODUCTION (Supabase)' if use_supabase else 'LOCAL (file)'}")
    
    sb_client = get_supabase_client()
    if not sb_client:
        return
    standings = generate_excel(
        sb_client, output_file, teams_abbr, rosters, capt_vc,
        match_history, all_match_nums, player_country, player_status=player_status, use_supabase=use_supabase
    )

    print("\n🏆 Current Standings:")
    for i, (team, abbr, pts, _) in enumerate(
        sorted(standings, key=lambda x: -x[2]), 1
    ):
        print(f"   {i:>2}. [{abbr:<3}] {team:<28} {pts:>8.1f} pts")

    print(f"\n✅ Done! Output: {output_file}\n")


def main():
    p = argparse.ArgumentParser(
        description="WWC Fantasy League — fetch match & regenerate standings Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python dwl_scoring_pipeline.py -m 1527674 -n 1
  python dwl_scoring_pipeline.py -m 1527675 -n 2
  python dwl_scoring_pipeline.py --force

Required Excel structure (single file):
  - Sheet "DWLTeams": Columns: Abbrv, Team Name
  - Sheet "Players": Columns: Player Name, Role, DWL Team, Country
  - Team sheets (e.g., "GD", "JR", etc.): Captain/VC info
        """,
    )
    p.add_argument("-m", "--match-id",  type=int, default=None,
                   help="ESPNcricinfo match ID (e.g. 1527674)")
    p.add_argument("-n", "--match-num", type=int, default=None,
                   help="Your league's match number for this game")
    p.add_argument("--config",    default="WWC_Config.xlsx",
                   help="Excel config file (default: WWC_Config.xlsx)")
    p.add_argument("-o", "--output",    default="DWL_Scores.xlsx",
                   help="Output Excel file (default: DWL_Scores.xlsx)")
    p.add_argument("--force",     action="store_true",
                   help="Overwrite existing match data")
    p.add_argument("-q", "--quiet",     action="store_true",
                   help="Skip printing the full scorecard")
    p.add_argument("-d", "--debug",     action="store_true",
                   help="Show debug output from the fetcher")

    args = p.parse_args()

    if args.match_id and args.match_num is None:
        p.error("-n/--match-num is required when -m/--match-id is given")

    asyncio.run(run(args))


if __name__ == "__main__":
    exit(main())